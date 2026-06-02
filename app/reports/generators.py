"""
Prospect BarberHub — Report generators
Cada builder retorna um dict normalizado; to_excel/to_pdf consomem esse dict.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime


# ── Utilitários ───────────────────────────────────────────────────────────────

def _fmt_brl(v: float | int) -> str:
    return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _period_str(d_from: date, d_to: date) -> str:
    return f"Período: {d_from.strftime('%d/%m/%Y')} a {d_to.strftime('%d/%m/%Y')}"


# ── Data builders ──────────────────────────────────────────────────────────────
# Cada função retorna um "report dict" com as chaves:
#   title, subtitle, sheet_name, filename, headers, rows, totals,
#   money_cols (0-indexed), col_widths (relativos), landscape

def build_appointments(date_from: date, date_to: date) -> dict:
    from app.models.appointment import Appointment

    appts = (
        Appointment.query
        .filter(
            Appointment.scheduled_date >= date_from,
            Appointment.scheduled_date <= date_to,
        )
        .order_by(Appointment.scheduled_date, Appointment.scheduled_time)
        .all()
    )

    headers = ["Data", "Hora", "Cliente", "Telefone", "Barbeiro", "Serviço", "Valor (R$)", "Status"]
    rows: list[list] = []
    total_revenue = 0.0

    for a in appts:
        price = float(a.service.price) if a.service else 0.0
        if a.status == "completed":
            total_revenue += price
        rows.append([
            a.scheduled_date.strftime("%d/%m/%Y"),
            a.scheduled_time.strftime("%H:%M"),
            a.customer.name if a.customer else "—",
            a.customer.phone or "—" if a.customer else "—",
            a.barber.name if a.barber else "—",
            a.service.name if a.service else "—",
            price,
            a.status_label,
        ])

    return {
        "title":      "Relatório de Agendamentos",
        "subtitle":   f"{_period_str(date_from, date_to)} · {len(rows)} registro(s)",
        "sheet_name": "Agendamentos",
        "filename":   f"agendamentos_{date_from}_{date_to}",
        "headers":    headers,
        "rows":       rows,
        "totals":     ["TOTAL", "", str(len(rows)) + " agend.", "", "", "", total_revenue, ""],
        "money_cols": [6],
        "col_widths": [2, 1.2, 3, 2.2, 2.5, 2.5, 2, 2.3],
        "landscape":  True,
    }


def build_revenue(date_from: date, date_to: date) -> dict:
    from app.models.appointment import Appointment

    appts = (
        Appointment.query
        .filter(
            Appointment.scheduled_date >= date_from,
            Appointment.scheduled_date <= date_to,
        )
        .order_by(Appointment.scheduled_date)
        .all()
    )

    # Agrupa por data em Python — evita SQL complexo com CASE no SQLite
    daily: dict[date, dict] = defaultdict(lambda: {
        "total": 0, "completed": 0, "cancelled": 0,
        "no_show": 0, "pending": 0, "revenue": 0.0,
    })
    for a in appts:
        d = daily[a.scheduled_date]
        d["total"] += 1
        if a.status == "completed":
            d["completed"] += 1
            if a.service:
                d["revenue"] += float(a.service.price)
        elif a.status == "cancelled":
            d["cancelled"] += 1
        elif a.status == "no_show":
            d["no_show"] += 1
        else:
            d["pending"] += 1

    headers = ["Data", "Total", "Concluídos", "Cancelados", "Não compareceu", "Em aberto", "Faturamento (R$)"]
    rows: list[list] = []
    tot = defaultdict(int)
    tot_rev = 0.0

    for dt in sorted(daily):
        d = daily[dt]
        rows.append([
            dt.strftime("%d/%m/%Y"),
            d["total"], d["completed"], d["cancelled"],
            d["no_show"], d["pending"], d["revenue"],
        ])
        for k in ("total", "completed", "cancelled", "no_show", "pending"):
            tot[k] += d[k]
        tot_rev += d["revenue"]

    return {
        "title":      "Relatório de Faturamento",
        "subtitle":   f"{_period_str(date_from, date_to)} · {len(rows)} dia(s) com atendimento",
        "sheet_name": "Faturamento",
        "filename":   f"faturamento_{date_from}_{date_to}",
        "headers":    headers,
        "rows":       rows,
        "totals":     [
            "TOTAL",
            tot["total"], tot["completed"], tot["cancelled"],
            tot["no_show"], tot["pending"], tot_rev,
        ],
        "money_cols": [6],
        "col_widths": [2, 1.5, 2, 2, 2.5, 2, 3],
        "landscape":  False,
    }


def build_services(date_from: date, date_to: date) -> dict:
    from app.models.appointment import Appointment

    appts = (
        Appointment.query
        .filter(
            Appointment.status == "completed",
            Appointment.scheduled_date >= date_from,
            Appointment.scheduled_date <= date_to,
        )
        .all()
    )

    svc: dict[int, dict] = defaultdict(lambda: {
        "name": "", "duration": 0, "price": 0.0, "count": 0, "revenue": 0.0,
    })
    for a in appts:
        if not a.service:
            continue
        s = svc[a.service_id]
        s["name"]     = a.service.name
        s["duration"] = a.service.duration_minutes
        s["price"]    = float(a.service.price)
        s["count"]   += 1
        s["revenue"] += float(a.service.price)

    ordered = sorted(svc.values(), key=lambda x: x["count"], reverse=True)
    tot_count   = sum(r["count"] for r in ordered)
    tot_revenue = sum(r["revenue"] for r in ordered)

    headers = ["Serviço", "Duração (min)", "Preço unit. (R$)", "Realizações", "Faturamento (R$)"]
    rows = [
        [r["name"], r["duration"], r["price"], r["count"], r["revenue"]]
        for r in ordered
    ]

    return {
        "title":      "Relatório de Serviços",
        "subtitle":   f"{_period_str(date_from, date_to)} · {tot_count} atendimentos concluídos",
        "sheet_name": "Serviços",
        "filename":   f"servicos_{date_from}_{date_to}",
        "headers":    headers,
        "rows":       rows,
        "totals":     ["TOTAL", "", "", tot_count, tot_revenue],
        "money_cols": [2, 4],
        "col_widths": [4, 2, 2.5, 2.5, 3],
        "landscape":  False,
    }


def build_barbers(date_from: date, date_to: date) -> dict:
    from app.models.appointment import Appointment

    appts = (
        Appointment.query
        .filter(
            Appointment.scheduled_date >= date_from,
            Appointment.scheduled_date <= date_to,
        )
        .all()
    )

    barbers: dict[int, dict] = defaultdict(lambda: {
        "name": "", "total": 0, "completed": 0,
        "cancelled": 0, "no_show": 0, "revenue": 0.0,
    })
    for a in appts:
        if not a.barber:
            continue
        b = barbers[a.barber_id]
        b["name"]  = a.barber.name
        b["total"] += 1
        if a.status == "completed":
            b["completed"] += 1
            if a.service:
                b["revenue"] += float(a.service.price)
        elif a.status == "cancelled":
            b["cancelled"] += 1
        elif a.status == "no_show":
            b["no_show"] += 1

    ordered = sorted(barbers.values(), key=lambda x: x["completed"], reverse=True)

    headers = [
        "Barbeiro", "Total", "Concluídos",
        "Cancelados", "Não compareceu", "Taxa conclusão", "Faturamento (R$)",
    ]
    rows: list[list] = []
    for r in ordered:
        closed = r["completed"] + r["cancelled"] + r["no_show"]
        rate   = f"{r['completed'] / closed * 100:.1f}%" if closed else "—"
        rows.append([
            r["name"], r["total"], r["completed"],
            r["cancelled"], r["no_show"], rate, r["revenue"],
        ])

    tot_total     = sum(r["total"] for r in ordered)
    tot_completed = sum(r["completed"] for r in ordered)
    tot_cancelled = sum(r["cancelled"] for r in ordered)
    tot_no_show   = sum(r["no_show"] for r in ordered)
    tot_revenue   = sum(r["revenue"] for r in ordered)
    tot_closed    = tot_completed + tot_cancelled + tot_no_show
    tot_rate      = f"{tot_completed / tot_closed * 100:.1f}%" if tot_closed else "—"

    return {
        "title":      "Relatório de Barbeiros",
        "subtitle":   f"{_period_str(date_from, date_to)} · {tot_completed} atendimentos concluídos",
        "sheet_name": "Barbeiros",
        "filename":   f"barbeiros_{date_from}_{date_to}",
        "headers":    headers,
        "rows":       rows,
        "totals":     [
            "TOTAL", tot_total, tot_completed,
            tot_cancelled, tot_no_show, tot_rate, tot_revenue,
        ],
        "money_cols": [6],
        "col_widths": [3, 1.5, 2, 2, 2.5, 2.5, 3],
        "landscape":  True,
    }


# ── Excel generator ────────────────────────────────────────────────────────────

def to_excel(report: dict) -> io.BytesIO:
    """
    Gera .xlsx estilizado a partir de um report dict.
    Levanta ImportError se openpyxl não estiver instalado.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report["sheet_name"][:31]

    # Paleta
    DARK_FILL  = PatternFill("solid", fgColor="1F2937")
    GOLD_FILL  = PatternFill("solid", fgColor="D4A017")
    ALT_FILL   = PatternFill("solid", fgColor="F3F4F6")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    thin_gray  = Side(style="thin", color="E5E7EB")
    grid_bdr   = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    n_cols   = len(report["headers"])
    last_col = get_column_letter(n_cols)

    def _merge_header_cell(row: int, value: str, font_kw: dict, fill, height: int, align="center"):
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws[f"A{row}"]
        c.value     = value
        c.font      = Font(**font_kw)
        c.fill      = fill
        c.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[row].height = height

    # Linha 1 — Título
    _merge_header_cell(
        1, f"Prospect BarberHub  ·  {report['title']}",
        {"bold": True, "size": 13, "color": "FFFFFF", "name": "Calibri"},
        DARK_FILL, 30,
    )
    # Linha 2 — Subtítulo
    _merge_header_cell(
        2, report["subtitle"],
        {"size": 9, "color": "6B7280", "name": "Calibri"},
        WHITE_FILL, 16,
    )
    # Linha 3 — Gerado em
    _merge_header_cell(
        3, f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        {"size": 8, "italic": True, "color": "9CA3AF", "name": "Calibri"},
        WHITE_FILL, 14, align="right",
    )
    ws.row_dimensions[4].height = 6  # espaçador

    # Linha 5 — Cabeçalho da tabela
    money_1idx = {c + 1 for c in report.get("money_cols", [])}
    for col_i, hdr in enumerate(report["headers"], start=1):
        c = ws.cell(row=5, column=col_i, value=hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill      = GOLD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(bottom=Side(style="medium", color="B8860B"))
    ws.row_dimensions[5].height = 22

    # Linhas de dados (a partir da 6)
    for r_i, row in enumerate(report["rows"], start=6):
        fill = ALT_FILL if r_i % 2 == 0 else WHITE_FILL
        for col_i, val in enumerate(row, start=1):
            c = ws.cell(row=r_i, column=col_i, value=val)
            c.fill   = fill
            c.font   = Font(size=10, name="Calibri")
            c.border = grid_bdr
            if col_i in money_1idx:
                c.number_format = '#,##0.00'
                c.alignment     = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r_i].height = 18

    # Linha de totais
    totals = report.get("totals")
    if totals:
        t_row = 6 + len(report["rows"])
        for col_i, val in enumerate(totals, start=1):
            c = ws.cell(row=t_row, column=col_i, value=val)
            c.fill   = GOLD_FILL
            c.font   = Font(bold=True, size=10, color="1F2937", name="Calibri")
            c.border = Border(top=Side(style="medium", color="B8860B"))
            if col_i in money_1idx and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
                c.alignment     = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[t_row].height = 20

    # Largura automática das colunas
    for col_i in range(1, n_cols + 1):
        col_letter = get_column_letter(col_i)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_i).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    ws.freeze_panes = "A6"  # congela cabeçalho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF generator ──────────────────────────────────────────────────────────────

def to_pdf(report: dict) -> io.BytesIO:
    """
    Gera .pdf profissional com ReportLab a partir de um report dict.
    Levanta ImportError se reportlab não estiver instalado.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.pagesizes import landscape as rl_landscape
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    # Cores
    C_DARK  = HexColor("#1F2937")
    C_GOLD  = HexColor("#D4A017")
    C_GOLD2 = HexColor("#B8860B")
    C_ALT   = HexColor("#F3F4F6")
    C_GRAY  = HexColor("#6B7280")
    C_LGRAY = HexColor("#E5E7EB")
    C_WHITE = colors.white

    buf     = io.BytesIO()
    margin  = 1.5 * cm
    pgsz    = rl_landscape(A4) if report.get("landscape") else A4
    avail_w = pgsz[0] - 2 * margin

    doc = SimpleDocTemplate(
        buf, pagesize=pgsz,
        rightMargin=margin, leftMargin=margin,
        topMargin=2 * cm, bottomMargin=1.5 * cm,
    )

    def _ps(name, **kw) -> ParagraphStyle:
        return ParagraphStyle(name, **kw)

    story: list = []

    # Cabeçalho textual
    story.append(Paragraph("Prospect BarberHub",
        _ps("brand", fontSize=16, fontName="Helvetica-Bold",
            textColor=C_GOLD, spaceAfter=2)))
    story.append(Paragraph(report["title"],
        _ps("title", fontSize=13, fontName="Helvetica-Bold",
            textColor=C_DARK, spaceAfter=2)))
    story.append(Paragraph(report["subtitle"],
        _ps("sub", fontSize=9, fontName="Helvetica",
            textColor=C_GRAY, spaceAfter=2)))
    story.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        _ps("gen", fontSize=7, fontName="Helvetica",
            textColor=C_GRAY, spaceAfter=6, alignment=TA_RIGHT)))
    story.append(HRFlowable(
        width="100%", thickness=1.5, color=C_GOLD,
        spaceBefore=2, spaceAfter=10))

    # Formata células: float → BRL, outros → str
    money_0idx = set(report.get("money_cols", []))

    def _cell(val, col_i: int) -> str:
        if val is None or val == "":
            return ""
        if col_i in money_0idx and isinstance(val, (int, float)):
            return _fmt_brl(val)
        return str(val)

    # Monta dados da tabela
    tbl_data = [report["headers"]]
    for row in report["rows"]:
        tbl_data.append([_cell(v, i) for i, v in enumerate(row)])

    totals = report.get("totals")
    if totals:
        tbl_data.append([_cell(v, i) for i, v in enumerate(totals)])

    # Largura das colunas
    raw_ws = report.get("col_widths")
    if raw_ws:
        s = sum(raw_ws)
        col_widths = [w / s * avail_w for w in raw_ws]
    else:
        col_widths = None

    n_data = len(report["rows"])

    # Estilo da tabela
    cmds = [
        # Cabeçalho
        ("BACKGROUND",    (0, 0), (-1, 0), C_DARK),
        ("TEXTCOLOR",     (0, 0), (-1, 0), C_WHITE),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 9),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        # Dados
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("ALIGN",         (0, 1), (-1, -1), "LEFT"),
        ("TOPPADDING",    (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        # Grade
        ("GRID",          (0, 0), (-1, -1), 0.3, C_LGRAY),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, C_GOLD),
    ]

    # Linhas alternadas
    for i in range(1, n_data + 1):
        if i % 2 == 0:
            cmds.append(("BACKGROUND", (0, i), (-1, i), C_ALT))

    # Colunas de valor: alinhamento à direita
    for col_i in money_0idx:
        cmds.append(("ALIGN", (col_i, 1), (col_i, -1), "RIGHT"))

    # Linha de totais
    if totals:
        cmds += [
            ("BACKGROUND", (0, -1), (-1, -1), C_GOLD),
            ("TEXTCOLOR",  (0, -1), (-1, -1), C_DARK),
            ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",   (0, -1), (-1, -1), 9),
            ("LINEABOVE",  (0, -1), (-1, -1), 1.5, C_GOLD2),
            ("ALIGN",      (0, -1), (-1, -1), "CENTER"),
        ]
        # Re-alinha colunas monetárias na linha de totais
        for col_i in money_0idx:
            cmds.append(("ALIGN", (col_i, -1), (col_i, -1), "RIGHT"))

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(cmds))
    story.append(tbl)

    # Rodapé
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Prospect BarberHub · Relatório gerado automaticamente pelo sistema",
        _ps("footer", fontSize=7, textColor=C_GRAY, alignment=TA_CENTER),
    ))

    doc.build(story)
    buf.seek(0)
    return buf
